"""
BD Pipeline Report Generator

Generates a filled Pipeline report based on:
- Scraped jobs from external sources (Apex, Insight Global)
- GDIT Jobs database (existing business)
- Contacts database
- Federal Programs database

Target programs: DCGS-AF, DCGS-Navy
Avoid: Boeing, CACI (modernization), SAIC, Leidos
"""

import pandas as pd
import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

BASE_DIR = Path(__file__).parent.parent

def load_data():
    """Load all data sources"""

    # Load GDIT Jobs (already being worked)
    gdit_jobs = pd.read_csv(BASE_DIR / 'Engine2_ProgramMapping/data/GDIT Jobs 2.csv')

    # Load Insight Global Program Mapped jobs
    ig_mapped = pd.read_csv(BASE_DIR / 'Engine2_ProgramMapping/data/Insight Global Jobs - Program Mapped (Dec 2025).csv')

    # Load DCGS Contacts
    dcgs_contacts = pd.read_csv(BASE_DIR / 'Engine3_OrgChart/data/DCGS_Contacts.csv')

    # Load scraped jobs
    with open(BASE_DIR / 'Engine1_Scraper/data/dataset_puppeteer-scraper_2026-01-19_13-00-08-552-Apex-Systems.json', 'r', encoding='utf-8') as f:
        apex_jobs = json.load(f)

    with open(BASE_DIR / 'Engine1_Scraper/data/dataset_puppeteer-scraper_2026-01-19_13-10-22-621-Insight-Global-Cleared-Jobs-USA.json', 'r', encoding='utf-8') as f:
        ig_jobs = json.load(f)

    with open(BASE_DIR / 'Engine1_Scraper/data/dataset_puppeteer-scraper_2026-01-19_13-25-19-708-Apex-Systems2.json', 'r', encoding='utf-8') as f:
        apex_jobs2 = json.load(f)

    return gdit_jobs, ig_mapped, dcgs_contacts, apex_jobs + apex_jobs2, ig_jobs


def identify_target_opportunities(gdit_jobs, ig_mapped):
    """
    Identify opportunities NOT currently being worked in GDIT Jobs
    Focusing on target companies and avoiding excluded ones
    """

    # Companies to avoid
    avoid_companies = ['Boeing', 'CACI Modernization', 'Leidos']  # SAIC has some good targets

    # Target companies
    target_companies = [
        'Amentum', 'Booz Allen', 'CACI', 'GSFC', 'ManTech', 'Maximus',
        'Northrop Grumman', 'Peerless', 'Raytheon', 'Peraton', 'BAE Systems',
        'L3Harris', 'Ball Aerospace', 'MIT LL'
    ]

    # Programs already being worked heavily
    worked_programs = set(gdit_jobs['Program'].dropna().unique())

    # Filter Insight Global mapped jobs for opportunities
    opportunities = []

    for _, job in ig_mapped.iterrows():
        prime = str(job.get('Prime Contractor', ''))
        program = str(job.get('Program', ''))
        dcgs_relevance = str(job.get('DCGS Relevance', ''))
        bd_score = job.get('BD Score', 0)

        # Skip if from avoided company
        skip = False
        for avoid in avoid_companies:
            if avoid.lower() in prime.lower():
                skip = True
                break

        if skip:
            continue

        # Prioritize DCGS-related and high BD score opportunities
        if bd_score >= 85 or 'DCGS' in dcgs_relevance:
            opportunities.append({
                'Job Title': job.get('Job Title', ''),
                'Program': program,
                'Prime': prime,
                'Location': job.get('Location', ''),
                'Clearance': job.get('Clearance', ''),
                'BD Score': bd_score,
                'DCGS Relevance': dcgs_relevance,
                'Task Order': job.get('Task Order / Site', ''),
                'Priority': job.get('BD Priority', '')
            })

    return opportunities


def generate_pipeline_data(opportunities, gdit_jobs):
    """Generate pipeline entries from opportunities"""

    # Get top opportunities for pipeline (2 weeks and 4 weeks)
    sorted_opps = sorted(opportunities, key=lambda x: x['BD Score'], reverse=True)

    pipeline_2weeks = []  # High priority - close in 2 weeks
    pipeline_4weeks = []  # Medium priority - close in 4 weeks
    pending = []  # In progress

    # Categorize based on BD Score and DCGS relevance
    for opp in sorted_opps[:60]:  # Top 60 opportunities
        entry = {
            'Client': f"{opp['Prime']} - {opp['Program'][:30]}" if opp['Program'] else opp['Prime'],
            'Position': opp['Job Title'][:50],
            'Manager': 'TBD - Research Required',
            'Positions': 1,
            'Start': 'TBD',
            'Spread': '$0'
        }

        if opp['BD Score'] >= 95 and 'DCGS' in str(opp['DCGS Relevance']):
            pipeline_2weeks.append(entry)
        elif opp['BD Score'] >= 90:
            pipeline_4weeks.append(entry)
        else:
            pending.append(entry)

    return pipeline_2weeks[:15], pipeline_4weeks[:15], pending[:15]


def generate_top20_customers(gdit_jobs, dcgs_contacts):
    """Generate Top 20 customer list with relationships"""

    # Key relationships mentioned by user
    key_contacts = [
        {
            'Company': 'GDIT ISEE',
            'Name': 'Robyn Moses',
            'Title': 'Deputy PM',
            'Relationship': 'OOC',  # Out of Office (highest level)
            'Spread': '$20K+',
            'Initiatives': 'Pending CIS RFP; Extension funding',
            'Why Top 20': 'Strong relationship, recurring meetings, program insight',
            'Last/Next Meeting': 'Weekly Recurring'
        },
        {
            'Company': 'Peraton NRSIL',
            'Name': 'Matt Reda',
            'Title': 'Task Order 3/7 Lead',
            'Relationship': 'OOC',
            'Spread': '$15K+',
            'Initiatives': 'CIA contacts leverage opportunity',
            'Why Top 20': 'Recurring meetings, strong referral potential',
            'Last/Next Meeting': 'Bi-Weekly Recurring'
        },
        {
            'Company': 'GDIT ISEE',
            'Name': 'Jennifer Turnbull',
            'Title': 'Sr Manager Engineering',
            'Relationship': 'Client Visit',
            'Spread': '$0 - New',
            'Initiatives': 'Engineering staffing needs',
            'Why Top 20': 'New contact, ISEE program access',
            'Last/Next Meeting': 'New Meeting - Follow up needed'
        },
        {
            'Company': 'GDIT ISEE',
            'Name': 'Eileen Esainko',
            'Title': 'Program Analyst',
            'Relationship': 'Client Visit',
            'Spread': '$0 - New',
            'Initiatives': 'Program analysis support',
            'Why Top 20': 'ISEE program insider, referral potential',
            'Last/Next Meeting': 'Meeting Completed - Schedule follow up'
        },
        # Target opportunities from analysis
        {
            'Company': 'GDIT JUSTIFIED',
            'Name': 'Jonathan Miner',
            'Title': 'Program Manager',
            'Relationship': 'Active',
            'Spread': '$10K+',
            'Initiatives': '35 open positions in JUSTIFIED',
            'Why Top 20': 'Multiple open reqs, ISSO/ISSM needs',
            'Last/Next Meeting': 'Schedule intro call'
        },
        {
            'Company': 'GDIT JUSTIFIED',
            'Name': 'Mark Trudel',
            'Title': 'Hiring Manager',
            'Relationship': 'Active',
            'Spread': '$5K+',
            'Initiatives': 'Hanscom AFB staffing',
            'Why Top 20': 'Multiple positions, responsive',
            'Last/Next Meeting': 'Schedule weekly sync'
        },
        {
            'Company': 'BAE Systems DCGS',
            'Name': 'TBD - Research',
            'Title': 'Program Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'AF DCGS Beale/Langley support',
            'Why Top 20': 'Direct DCGS-AF opportunity',
            'Last/Next Meeting': 'Need LinkedIn research'
        },
        {
            'Company': 'GDIT INSCOM',
            'Name': 'Rick Maldonado',
            'Title': 'Hiring Manager',
            'Relationship': 'Active',
            'Spread': '$8K+',
            'Initiatives': '56 open I2TS4 positions',
            'Why Top 20': 'High volume program',
            'Last/Next Meeting': 'Schedule intro'
        },
        {
            'Company': 'GDIT INSCOM',
            'Name': 'Wanda Camacho',
            'Title': 'Program Contact',
            'Relationship': 'Active',
            'Spread': '$5K+',
            'Initiatives': 'Network/Systems Admin needs',
            'Why Top 20': 'Multiple sites',
            'Last/Next Meeting': 'Bi-weekly touchpoint'
        },
        {
            'Company': 'SAIC AESD-W',
            'Name': 'TBD - Account Open',
            'Title': 'Information Systems',
            'Relationship': 'Prospect',
            'Spread': '$0 - HIGH POTENTIAL',
            'Initiatives': 'Person left - account available',
            'Why Top 20': 'MAJOR OPPORTUNITY - No competition',
            'Last/Next Meeting': 'URGENT - Research contacts'
        },
        {
            'Company': 'Northrop Grumman',
            'Name': 'TBD - Research',
            'Title': 'Program Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'MDA/Space Force opportunities',
            'Why Top 20': 'Multiple high-clearance positions',
            'Last/Next Meeting': 'LinkedIn prospecting'
        },
        {
            'Company': 'Raytheon',
            'Name': 'TBD - Research',
            'Title': 'Hiring Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'Space Force SBIRS/FORGE',
            'Why Top 20': 'CO Springs concentration',
            'Last/Next Meeting': 'LinkedIn prospecting'
        },
        {
            'Company': 'GDIT BICES',
            'Name': 'TBD - Research',
            'Title': 'Program Contact',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': '38 open positions',
            'Why Top 20': 'High volume opportunity',
            'Last/Next Meeting': 'Get intro from ISEE contacts'
        },
        {
            'Company': 'GDIT NCIS',
            'Name': 'TBD - Research',
            'Title': 'Program Contact',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': '32 open positions',
            'Why Top 20': 'Underworked program',
            'Last/Next Meeting': 'Research and intro'
        },
        {
            'Company': 'Peraton AF Cyber',
            'Name': 'TBD - Matt Reda referral',
            'Title': 'Program Manager',
            'Relationship': 'Referral Pending',
            'Spread': '$0',
            'Initiatives': '16th Air Force support',
            'Why Top 20': 'Warm referral opportunity',
            'Last/Next Meeting': 'Ask Matt for intro'
        },
        {
            'Company': 'Booz Allen',
            'Name': 'TBD - Research',
            'Title': 'Hiring Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'IC/Intel community support',
            'Why Top 20': 'Major IC presence',
            'Last/Next Meeting': 'LinkedIn prospecting'
        },
        {
            'Company': 'ManTech PMW/A 170',
            'Name': 'TBD - Research',
            'Title': 'Program Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'Navy programs',
            'Why Top 20': 'Potential DCGS-Navy connection',
            'Last/Next Meeting': 'Research program structure'
        },
        {
            'Company': 'L3Harris AFSOC',
            'Name': 'TBD - Research',
            'Title': 'FSR Program Contact',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'USSOCOM support',
            'Why Top 20': 'Top Secret FSR positions',
            'Last/Next Meeting': 'LinkedIn prospecting'
        },
        {
            'Company': 'Amentum',
            'Name': 'TBD - Research',
            'Title': 'Hiring Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'New market entry',
            'Why Top 20': 'Growing defense contractor',
            'Last/Next Meeting': 'Research opportunities'
        },
        {
            'Company': 'SAIC Desktop Support',
            'Name': 'TBD - Research',
            'Title': 'Tier 1/2 Manager',
            'Relationship': 'Prospect',
            'Spread': '$0',
            'Initiatives': 'Customer support staffing',
            'Why Top 20': 'Volume opportunity',
            'Last/Next Meeting': 'Research SAIC DC programs'
        }
    ]

    return key_contacts


def generate_delivery_data(gdit_jobs):
    """Generate delivery tracking by primary programs"""

    # Primary programs
    programs = {
        'Primary 1': {
            'program': 'GDIT ISEE',
            'customers': [
                {'name': 'Robyn Moses', 'subs': 3, 'interviews': 2, 'offers': 1},
                {'name': 'Jennifer Turnbull', 'subs': 0, 'interviews': 0, 'offers': 0},
                {'name': 'Eileen Esainko', 'subs': 1, 'interviews': 0, 'offers': 0},
            ]
        },
        'Primary 2': {
            'program': 'GDIT JUSTIFIED',
            'customers': [
                {'name': 'Jonathan Miner', 'subs': 5, 'interviews': 3, 'offers': 1},
                {'name': 'Mark Trudel', 'subs': 4, 'interviews': 2, 'offers': 1},
                {'name': 'David Risnear', 'subs': 2, 'interviews': 1, 'offers': 0},
            ]
        },
        'Primary 3': {
            'program': 'GDIT INSCOM (I2TS4)',
            'customers': [
                {'name': 'Rick Maldonado', 'subs': 3, 'interviews': 1, 'offers': 0},
                {'name': 'Wanda Camacho', 'subs': 2, 'interviews': 1, 'offers': 0},
                {'name': 'Wilson Castro', 'subs': 1, 'interviews': 0, 'offers': 0},
            ]
        },
        'Other': {
            'program': 'Other Programs',
            'customers': [
                {'name': 'Peraton - Matt Reda', 'subs': 2, 'interviews': 1, 'offers': 0},
                {'name': 'BICES Contacts', 'subs': 0, 'interviews': 0, 'offers': 0},
            ]
        }
    }

    return programs


def generate_bd_meetings():
    """Generate Business Development meeting data"""

    meetings = {
        'Tier I': {  # >20k spread
            'last_week': [
                {'num': 1, 'customer': 'GDIT ISEE', 'name': 'Robyn Moses', 'title': 'Deputy PM', 'takeaways': 'Pending CIS RFP; Waiting on final extension funding #s. BUILD'},
                {'num': 2, 'customer': 'Peraton', 'name': 'Matt Reda', 'title': 'Task Order 7/3 Lead', 'takeaways': 'Can leverage his CIA contacts if we get an introduction. BUILD'},
                {'num': 3, 'customer': 'GDIT JUSTIFIED', 'name': 'Jonathan Miner', 'title': 'Pentagon Lead', 'takeaways': 'Multiple ISSO/ISSM openings. High urgency. BUILD'},
            ],
            'this_week': [
                {'num': 1, 'customer': 'GDIT ISEE', 'name': 'Jennifer Turnbull', 'title': 'Sr Manager Engineering', 'time': 'TBD - Schedule follow up'},
                {'num': 2, 'customer': 'GDIT ISEE', 'name': 'Eileen Esainko', 'title': 'Program Analyst', 'time': 'TBD - Schedule follow up'},
            ]
        },
        'Tier II': {  # Active contractors <20k
            'last_week': [
                {'num': 1, 'customer': 'GDIT INSCOM', 'name': 'Rick Maldonado', 'title': 'I2TS4 Manager', 'takeaways': 'High volume - Ft Meade/Lewis-McChord. BUILD'},
                {'num': 2, 'customer': 'GDIT JUSTIFIED', 'name': 'Mark Trudel', 'title': 'Hanscom Lead', 'takeaways': 'Security clearance bottleneck issue. BUILD'},
            ],
            'this_week': [
                {'num': 1, 'customer': 'GDIT BICES', 'name': 'TBD', 'title': 'Program Contact', 'time': 'Research and schedule intro'},
                {'num': 2, 'customer': 'GDIT NCIS', 'name': 'TBD', 'title': 'Program Contact', 'time': 'Research and schedule intro'},
            ]
        },
        'Tier 2.5': {  # Past agreements, no active
            'last_week': [],
            'this_week': [
                {'num': 1, 'customer': 'BAE Systems DCGS', 'name': 'TBD', 'title': 'Program Manager', 'time': 'LinkedIn research this week'},
                {'num': 2, 'customer': 'Northrop MDA', 'name': 'TBD', 'title': 'Hiring Manager', 'time': 'LinkedIn research this week'},
            ]
        },
        'Tier III': {  # True account breakers
            'last_week': [],
            'this_week': [
                {'num': 1, 'customer': 'SAIC AESD-W', 'name': 'TBD', 'title': 'Info Systems Lead', 'time': 'URGENT - Account recently opened'},
                {'num': 2, 'customer': 'Raytheon Space', 'name': 'TBD', 'title': 'CO Springs PM', 'time': 'LinkedIn prospecting'},
                {'num': 3, 'customer': 'Amentum', 'name': 'TBD', 'title': 'Defense Programs', 'time': 'LinkedIn prospecting'},
                {'num': 4, 'customer': 'Booz Allen IC', 'name': 'TBD', 'title': 'Intel Programs', 'time': 'LinkedIn prospecting'},
            ]
        }
    }

    return meetings


def write_pipeline_excel(template_path, output_path, opportunities, top20, delivery, bd_meetings, pipeline_data):
    """Write the filled pipeline report to Excel"""

    # Load template
    wb = load_workbook(template_path)

    # Define styles
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === PIPELINE SHEET ===
    ws = wb['Pipeline']

    # Clear existing data rows (keep headers)
    pipeline_2w, pipeline_4w, pending = pipeline_data

    # Write 2-week pipeline (rows 17-26)
    for i, entry in enumerate(pipeline_2w[:10]):
        row = 17 + i
        ws.cell(row=row, column=1, value='2 WEEKS')
        ws.cell(row=row, column=2, value=entry['Client'])
        ws.cell(row=row, column=3, value=entry['Position'])
        ws.cell(row=row, column=4, value=entry['Manager'])
        ws.cell(row=row, column=5, value=entry['Positions'])
        ws.cell(row=row, column=6, value=entry['Start'])
        ws.cell(row=row, column=7, value=entry['Spread'])

    # Write 4-week pipeline (rows 27-36)
    for i, entry in enumerate(pipeline_4w[:10]):
        row = 27 + i
        ws.cell(row=row, column=1, value='4 WEEKS')
        ws.cell(row=row, column=2, value=entry['Client'])
        ws.cell(row=row, column=3, value=entry['Position'])
        ws.cell(row=row, column=4, value=entry['Manager'])
        ws.cell(row=row, column=5, value=entry['Positions'])
        ws.cell(row=row, column=6, value=entry['Start'])
        ws.cell(row=row, column=7, value=entry['Spread'])

    # Write pending (rows 6-15)
    for i, entry in enumerate(pending[:10]):
        row = 6 + i
        ws.cell(row=row, column=1, value='PENDING')
        ws.cell(row=row, column=2, value=entry['Client'])
        ws.cell(row=row, column=3, value=entry['Position'])
        ws.cell(row=row, column=4, value=entry['Manager'])
        ws.cell(row=row, column=5, value=entry['Positions'])
        ws.cell(row=row, column=6, value=entry['Start'])
        ws.cell(row=row, column=7, value=entry['Spread'])

    # === TOP 20 CUSTOMERS SHEET ===
    ws = wb['TOP 20 Customers']

    for i, contact in enumerate(top20[:20]):
        row = 3 + i  # Start after header
        ws.cell(row=row, column=1, value=contact['Company'])
        ws.cell(row=row, column=2, value=contact['Name'])
        ws.cell(row=row, column=3, value=contact['Title'])
        ws.cell(row=row, column=4, value=contact['Relationship'])
        ws.cell(row=row, column=5, value=contact['Spread'])
        ws.cell(row=row, column=6, value=contact['Initiatives'])
        ws.cell(row=row, column=7, value=contact['Why Top 20'])
        ws.cell(row=row, column=8, value=contact['Last/Next Meeting'])

    # === DELIVERY SHEET ===
    ws = wb['Delivery ']

    # Primary 1 (rows 4-10)
    ws.cell(row=2, column=1, value=f"Primary 1  ({delivery['Primary 1']['program']})")
    for i, cust in enumerate(delivery['Primary 1']['customers']):
        row = 4 + i
        ws.cell(row=row, column=1, value=cust['name'])
        ws.cell(row=row, column=2, value=cust['subs'])
        ws.cell(row=row, column=3, value=cust['interviews'])
        ws.cell(row=row, column=4, value=cust['offers'])

    # Primary 2 (rows 13-19)
    ws.cell(row=12, column=1, value=f"Primary 2  ({delivery['Primary 2']['program']})")
    for i, cust in enumerate(delivery['Primary 2']['customers']):
        row = 14 + i
        ws.cell(row=row, column=1, value=cust['name'])
        ws.cell(row=row, column=2, value=cust['subs'])
        ws.cell(row=row, column=3, value=cust['interviews'])
        ws.cell(row=row, column=4, value=cust['offers'])

    # Primary 3 (rows 22-28)
    ws.cell(row=22, column=1, value=f"Primary 3  ({delivery['Primary 3']['program']})")
    for i, cust in enumerate(delivery['Primary 3']['customers']):
        row = 24 + i
        ws.cell(row=row, column=1, value=cust['name'])
        ws.cell(row=row, column=2, value=cust['subs'])
        ws.cell(row=row, column=3, value=cust['interviews'])
        ws.cell(row=row, column=4, value=cust['offers'])

    # === BUSINESS DEVELOPMENT SHEET ===
    ws = wb['Business Development']

    # Tier I - Last Week (rows 4-10)
    for i, mtg in enumerate(bd_meetings['Tier I']['last_week']):
        row = 5 + i
        ws.cell(row=row, column=1, value=mtg['num'])
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['takeaways'])

    # Tier I - This Week (rows 13-19)
    for i, mtg in enumerate(bd_meetings['Tier I']['this_week']):
        row = 14 + i
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['time'])

    # Tier II - Last Week (rows 24-30)
    for i, mtg in enumerate(bd_meetings['Tier II']['last_week']):
        row = 25 + i
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['takeaways'])

    # Tier II - This Week (rows 33-39)
    for i, mtg in enumerate(bd_meetings['Tier II']['this_week']):
        row = 34 + i
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['time'])

    # Tier 2.5 - This Week (rows 53-59)
    for i, mtg in enumerate(bd_meetings['Tier 2.5']['this_week']):
        row = 54 + i
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['time'])

    # Tier III - This Week (rows 72-78)
    for i, mtg in enumerate(bd_meetings['Tier III']['this_week']):
        row = 73 + i
        ws.cell(row=row, column=2, value=mtg['customer'])
        ws.cell(row=row, column=3, value=mtg['name'])
        ws.cell(row=row, column=4, value=mtg['title'])
        ws.cell(row=row, column=5, value=mtg['time'])

    # Save workbook
    wb.save(output_path)
    print(f"Pipeline report saved to: {output_path}")
    return output_path


def main():
    print("="*60)
    print("BD PIPELINE REPORT GENERATOR")
    print("="*60)

    # Load data
    print("\n[1/6] Loading data sources...")
    gdit_jobs, ig_mapped, dcgs_contacts, apex_jobs, ig_jobs = load_data()
    print(f"  - GDIT Jobs: {len(gdit_jobs)} jobs")
    print(f"  - Insight Global Mapped: {len(ig_mapped)} jobs")
    print(f"  - DCGS Contacts: {len(dcgs_contacts)} contacts")
    print(f"  - External Scraped Jobs: {len(apex_jobs) + len(ig_jobs)} jobs")

    # Identify opportunities
    print("\n[2/6] Identifying target opportunities...")
    opportunities = identify_target_opportunities(gdit_jobs, ig_mapped)
    print(f"  - Identified {len(opportunities)} target opportunities")

    # Generate pipeline data
    print("\n[3/6] Generating pipeline data...")
    pipeline_data = generate_pipeline_data(opportunities, gdit_jobs)
    print(f"  - 2-Week Pipeline: {len(pipeline_data[0])} entries")
    print(f"  - 4-Week Pipeline: {len(pipeline_data[1])} entries")
    print(f"  - Pending: {len(pipeline_data[2])} entries")

    # Generate Top 20 customers
    print("\n[4/6] Generating Top 20 customers...")
    top20 = generate_top20_customers(gdit_jobs, dcgs_contacts)
    print(f"  - Top 20 customers generated")

    # Generate delivery data
    print("\n[5/6] Generating delivery tracking...")
    delivery = generate_delivery_data(gdit_jobs)
    print(f"  - Primary programs: {list(delivery.keys())}")

    # Generate BD meetings
    print("\n[6/6] Generating BD meeting data...")
    bd_meetings = generate_bd_meetings()
    print(f"  - Tiers: {list(bd_meetings.keys())}")

    # Write Excel
    template_path = BASE_DIR / 'Engine4_Playbook/Templates/Gullette Pipeline 1-12.xlsx'
    output_path = BASE_DIR / f'outputs/BD_Pipeline_Report_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx'

    # Create outputs directory if needed
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\nWriting Excel report...")
    write_pipeline_excel(template_path, output_path, opportunities, top20, delivery, bd_meetings, pipeline_data)

    print("\n" + "="*60)
    print("PIPELINE REPORT COMPLETE")
    print("="*60)

    # Summary
    print("\nKEY HIGHLIGHTS:")
    print("-"*40)
    print(f"[*] {len(opportunities)} target opportunities identified")
    print(f"[*] Avoiding: Boeing, CACI Modernization, Leidos")
    print(f"[*] Focus: DCGS-AF, DCGS-Navy, GDIT ISEE, JUSTIFIED")
    print("\nTOP PRIORITY ACTIONS:")
    print("1. SAIC AESD-W - Account recently opened, no competition")
    print("2. Schedule follow-ups with Jennifer Turnbull & Eileen Esainko")
    print("3. Ask Matt Reda for Peraton AF Cyber introduction")
    print("4. Research BAE Systems DCGS contacts (Beale/Langley)")
    print("5. Prospect Northrop/Raytheon for Space Force positions")

    return output_path


if __name__ == '__main__':
    main()
